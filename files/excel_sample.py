from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Color

# Load an Excel file
wb = load_workbook('test.xlsx')

# Select the active worksheet
ws = wb.active

# Print cell value
print(ws['A1'].value)



# Create an Excel file
wb = Workbook()

# Select the active worksheet
ws = wb.active

# Write values to cells
ws['A1'] = "Hello"
ws['B1'] = "World"

# Save the workbook
wb.save('test.xlsx')



# Create an Excel file
wb = Workbook()
ws = wb.active

# Append rows of data
data = [["Hello", "World"], ["Python", "Excel"], ["OpenAI", "GPT-4"]]
for row in data:
    ws.append(row)

# Save the workbook
wb.save('test.xlsx')



# Load an Excel file
wb = load_workbook('test.xlsx')
ws = wb.active

# Iterate through rows
for row in ws.iter_rows(values_only=True):
    print(row)



from openpyxl import Workbook

# Create an Excel file
wb = Workbook()

# Create a new sheet
ws = wb.create_sheet("New Sheet")

# Write to the new sheet
ws['A1'] = "New Sheet"



# Write values
ws['A1'] = 5
ws['A2'] = 3

# Add a formula
ws['A3'] = "=SUM(A1:A2)"



# Write to a cell
cell = ws['A1']
cell.value = "Hello, World!"

# Apply style
font = Font(color=Color(rgb="FFFFFF00"), bold=True)
cell.font = font

# Save the workbook
wb.save('test.xlsx')
